# -*- coding:utf8 -*-
import pdb
import re
#import xlwt
from openpyxl import Workbook
#pdb.set_trace()
value = re.compile(r'^[-+]?[0-9]+\.[0-9]+$')

InputFileName = "C:\\Users\\Administrator\\Desktop\\evalue\\v3.txt"
#InputFileName = '/home4/dbs/logfile/train/v3.txt'

OutputFileName= InputFileName[:-4]+".xlsx"
fin = open(InputFileName, "r")
lines = fin.readlines()
fin.close()
fout = open(OutputFileName,"w")
i = 1
k = 1
nold=-1
Excel = Workbook()
Table = Excel.active
#Tabl2 = Excel.create_sheet()
Table.title = "Table"
for line in lines:
	data = line.split()
	if len(data)>2 and value.match(data[2]) and float(data[2])<1100 and float(data[2])>0.01:
		NFlag = data[0].split(':')
		if len(NFlag)>1 and NFlag[0].isdigit():
			Num = int(NFlag[0])
			if 1==k : nold = Num-1
			if nold<Num and Num<nold+100:
				nold = Num
				Table.cell(k, 1).value=Num
				Table.cell(k, 2).value=float(data[2])
				k = k+1
			else:
				print Num
		else:
			print NFlag
	else:
		print i,line
	i = i+1

Excel.save(OutputFileName)

