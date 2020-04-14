# -*- coding:utf8 -*-
import os
import pdb
import re
import xlwt
from openpyxl import Workbook
#pdb.set_trace()

g_EvaluationResultFileName = '/home4/dbs/logfile/result_VOC_01.txt'

def readRecallPresitionUoI(vResultFileName):
	Fin = open(vResultFileName, 'r')
	Lines = Fin.readlines()
	Fin.close()
	Table = []
	for i in range(len(Lines)):		
		l = Lines[i].strip()
		if 0 == len(l):continue
		if l.startswith("The evaluating results"):
			WeightFileName = os.path.basename(l.split()[-1])
			Table.append([WeightFileName])
		elif l.startswith('The result from'):
			ValTestFileName = l.split()[-1]
			Table[-1].append(ValTestFileName)
		else:
			data = l.split()
			Recall = data[-4].split(':')[1][:-1]
			Precision = data[-1].split(':')[1][:-1]
			UoI = data[-5][:-1]
			Table[-1].append(Recall)
			Table[-1].append(Precision)
			Table[-1].append(UoI)
		
	return Table
	
def write2Excel(voFileName, vTable):
	Excel = Workbook()
	Table = Excel.active
	#Tabl2 = Excel.create_sheet()
	Table.title = "Evaluations"
	if len(vTable)==0:return
	First = vTable[0]
	NumSrc = (len(First)-1)/4
	Table.cell(2, 1).value = '迭代次数'
	for i in range(NumSrc):#表头
		Table.cell(1, 2+i*4).value = First[1+i*4]
		Table.cell(2, 2+i*4).value = 'Recall'
		Table.cell(2, 2+i*4+1).value = 'Precision'
		Table.cell(2, 2+i*4+2).value = 'UoI'
	for i in range(len(vTable)):
		w = vTable[i]
		Table.cell(3+i, 1).value = w[0]
		for k in range(NumSrc):
			for c in range(3):
				Table.cell(3+i, 2+k*4 +c ).value = float(w[2+k*4 +c])
	Excel.save(voFileName)
	
if __name__=="__main__":
	Table = readRecallPresitionUoI(g_EvaluationResultFileName)
	EvaluationsFileName = os.path.splitext(g_EvaluationResultFileName)[0]+'.xlsx'
	write2Excel(EvaluationsFileName, Table)